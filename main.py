from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, NoSuchElementException
from tkinter import simpledialog, messagebox, Toplevel, Label, Button
import tkinter as tk
import openpyxl
from openpyxl.styles import Font
import inferenceModel as ML
import os
import sys
import re
import threading

from time import sleep

API_KEY = 'your API key' #paste your Api key

# def update_button_text(dialog, button, remaining_seconds):
#     if remaining_seconds > 0:
#         button.config(text=f"OK ({remaining_seconds})")
#         dialog.after(1000, update_button_text, dialog, button, remaining_seconds - 1)
#     else:
#         dialog.destroy()

# def show_custom_messagebox():
#     dialog = Toplevel(root)
#     dialog.title("Error")
    
#     message = Label(dialog, text="Wrong password, click OK to continue with upcoming stores.")
#     message.pack(padx=20, pady=10)
    
#     ok_button = Button(dialog, text="OK (5)", command=dialog.destroy)
#     ok_button.pack(pady=(0, 10))
    
#     # Start the countdown
#     update_button_text(dialog, ok_button, 5)
    
#     dialog.grab_set()

def year(year):
    global Year
    Year = year

def update_api_key(new_api_key):
    global API_KEY
    API_KEY = new_api_key
    with open(__file__, 'r') as file:
        lines = file.readlines()
    
    with open(__file__, 'w') as file:
        for line in lines:
            if line.startswith("API_KEY"):
                file.write(f'API_KEY = "{new_api_key}"\n')
            else:
                file.write(line)

def click_element(browser, wait, locator, retries=3, wait_time=1):
    for attempt in range(retries):
        try:
            element = wait.until(EC.element_to_be_clickable(locator))
            element.click()
            return
        except ElementClickInterceptedException:
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "dimmer-holder")))
            except TimeoutException:
                pass
            sleep(wait_time)
        except TimeoutException:
            pass

    element = wait.until(EC.presence_of_element_located(locator))
    browser.execute_script("arguments[0].click();", element)

# Define the desired year, quarter, and month
def quarter(selected_option):
    global Months
    global Quarter
    Quarter = selected_option  # Desired quarter in the respective format
    if Quarter == "Quarter 1 (Apr - Jun)":
        Month_1 = "April"
        Month_2 = "May"
        Month_3 = "June"
    elif Quarter == "Quarter 2 (Jul - Sep)":
        Month_1 = "July"
        Month_2 = "August"
        Month_3 = "September"
    elif Quarter == "Quarter 3 (Oct - Dec)":
        Month_1 = "October"
        Month_2 = "November"
        Month_3 = "December"
    else:
        Month_1 = "January"
        Month_2 = "February"
        Month_3 = "March"
    Months = [Month_1, Month_2, Month_3]

def monthly(selected_option):
    global monthly_quarter
    global Month
    Month = selected_option  # Desired month in the respective format
    if Month in ["April", "May", "June"]:
        monthly_quarter = "Quarter 1 (Apr - Jun)"
    elif Month in ["July", "August", "September"]:
        monthly_quarter = "Quarter 2 (Jul - Sep)"
    elif Month in ["October", "November", "December"]:
        monthly_quarter = "Quarter 3 (Oct - Dec)"
    else:
        monthly_quarter = "Quarter 4 (Jan - Mar)"

def all():
    global all_months
    all_months = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]

def quarter_decide(i):
    global all_quarter
    if i in [0, 1, 2]:
        all_quarter = "Quarter 1 (Apr - Jun)"
    elif i in [3, 4, 5]:
        all_quarter = "Quarter 2 (Jul - Sep)"
    elif i in [6, 7, 8]:
        all_quarter = "Quarter 3 (Oct - Dec)"
    else:
        all_quarter = "Quarter 4 (Jan - Mar)"
    
# Set up Chrome options and initial preferences
options = webdriver.ChromeOptions()
prefs = {
    "download.prompt_for_download": False,  # Disable download prompts
}
options.add_experimental_option("prefs", prefs)

def wrong_pass(browser, workbook, row, sheet, download_path):
    """Function to handle a wrong password."""
    try:
        # Check for the wrong password element
        wrong_pass_element = browser.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/div/div/div/alert-message/div")
        if wrong_pass_element:
            # threading.Thread(target=show_custom_messagebox).start()
            sheet.cell(row=row, column=4).value = "❌"
            sheet.cell(row=row, column=5).value = "wrong password"
            workbook.save(path)
            delete_folder_if_empty(download_path)
            return True
    except Exception:
        return False  # No wrong password found, continue execution

def kyc_auth(browser, wait):
    """Function to handle KYC authentication prompt."""
    try:
        sleep(0.5)
        kyc_auth_element = browser.find_element(By.XPATH, "/html/body[@class='modal-open']/adhr-table/div[@id='adhrtableV']/div[@class='modal-dialog sweet']/div[@class='modal-content']/div[@class='modal-footer']/a[@class='btn btn-primary'][2]")
        if kyc_auth_element:
            remind_later_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body[@class='modal-open']/adhr-table/div[@id='adhrtableV']/div[@class='modal-dialog sweet']/div[@class='modal-content']/div[@class='modal-footer']/a[@class='btn btn-primary'][2]"))).click()
    except Exception:
        pass  # No KYC prompt found, continue execution

def change_pass_old(browser, row, wait, sheet, workbook):
    """Function to handle the scenario where the old password is used again."""
    try:
        change_pass_old_element = browser.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div/div/div[2]/div/form/div[4]/div/div/span[2]")
        if change_pass_old_element:
            messagebox.showinfo("Error", "New password cannot be same as last 3 passwords. Enter a new password to continue.")
            # root = tk.Tk()
            # root.withdraw()
            # root.after(0, root.lift)
            # root.focus_force()
            browser.find_element(By.ID, "fo-newpwd ").clear()
            browser.find_element(By.ID, "fo-repwd").clear()
            new_password = simpledialog.askstring("Password", "Enter new password", show='*')
            if new_password is None:
                browser.quit()
            else:
                sheet.cell(row=row, column=3).value = new_password
                sheet.cell(row=row, column=5).value = "changed password"
                cell = sheet.cell(row=row, column=5)
                cell.font = Font(color="FF0000")
                workbook.save(path)
                workbook.close()
                browser.find_element(By.ID, "fo-newpwd ").send_keys(new_password + Keys.TAB + new_password +Keys.ENTER)
                change_pass_old(browser, row, wait, sheet, workbook)
                root.after(1000)
                return True
    except Exception:
        pass  # No old password issue found, continue execution

def change_pass(browser, wait, row, sheet, password, name, workbook):
    """Function to handle password change prompt."""
    try:
        change_pass_element = browser.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div/div/div[1]/div/div[1]/span")
        if change_pass_element:
            messagebox.showinfo("Error", f"Password needs to be changed for {name}, Enter a new password to continue.")
            wait.until(EC.presence_of_element_located((By.ID, "fo-pwd"))).send_keys(password)
            # root = tk.Tk()
            # root.withdraw()
            # root.after(0, root.lift)
            # root.focus_force()
            new_password = simpledialog.askstring("Password", "Enter new password", show='*')
            if new_password is None:
                browser.quit()
            else:
                sheet.cell(row=row, column=3).value = new_password
                sheet.cell(row=row, column=5).value = "changed password"
                cell = sheet.cell(row=row, column=5)
                cell.font = Font(color="FF0000")
                workbook.save(path)
                workbook.close()
                browser.find_element(By.ID, "fo-newpwd ").send_keys(new_password + Keys.TAB + new_password +Keys.ENTER)
                change_pass_old(browser, row, wait, sheet, workbook)
                root.after(1000)
                return True
    except Exception:
        return False  # No password change prompt found, continue execution

def captcha_err(browser, wait, password, captcha_val):
    """Function to handle incorrect captcha entries and prompt the user to re-enter the captcha."""
    while True:
        try:
            # Check if there's a captcha error element
            captcha_error_element = browser.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/form/div[4]/div/span")
            if captcha_error_element:
                # Prompt the user to re-enter the captcha
                wait.until(EC.presence_of_element_located((By.NAME, "user_name"))).send_keys(Keys.TAB + password + Keys.TAB)
                # root = tk.Tk()
                # root.withdraw()
                # root.after(0, root.lift)
                # root.focus_force()
                if captcha_val == "manual":
                    captcha = simpledialog.askstring("Captcha", "Enter the captcha")
                    if captcha is None:
                        browser.quit()
                        return
                elif captcha_val == "auto":
                    if os.path.exists(f"{dwnld_path}/captcha.png"):
                        os.remove(f"{dwnld_path}/captcha.png")
                    sleep(0.5)
                    element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/form/div[5]/div/div/div/table/tbody/tr[1]/th[1]/img')))
                    element.screenshot(f"{dwnld_path}/captcha.png")
                    captcha = captch_solver(dwnld_path)
                elif captcha_val == "ML":
                    if os.path.exists(f"{dwnld_path}/captcha.png"):
                        os.remove(f"{dwnld_path}/captcha.png")
                    sleep(1)
                    element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/form/div[5]/div/div/div/table/tbody/tr[1]/th[1]/img')))
                    element.screenshot(f"{dwnld_path}/captcha.png")
                    mlpath = f"{dwnld_path}/captcha.png"
                    captcha = ML.inference(mlpath)
                captcha_input = wait.until(EC.presence_of_element_located((By.NAME, "captcha")))
                captcha_input.clear()  # Clear previous input
                captcha_input.send_keys(captcha + Keys.ENTER)
                # Wait for the page to respond
                root.after(1500)
            else:
                # If no captcha error, break the loop
                break
        except NoSuchElementException:
            # Handle NoSuchElementException if captcha error element is not found
            break
        except Exception as e:
            # Handle other exceptions
            messagebox.showinfo("Exception", f"An unexpected error occurred: {str(e)}")
            break

def captch_solver(dwnld_path):
    sys.path.append(os.path.dirname(os.path.dirname(os.path.realpath(__file__))))

    from twocaptcha import TwoCaptcha

    api_key = os.getenv('APIKEY_2CAPTCHA', API_KEY)

    solver = TwoCaptcha(api_key)

    try:
        result = solver.normal(f'{dwnld_path}/captcha.png')
        print(result['code'])
        return result['code']

    except Exception as e:
        return str(e)

def delete_folder_if_empty(download_path):
    if os.path.exists(download_path) and os.path.isdir(download_path):
        if not os.listdir(download_path):
            os.rmdir(download_path)
            return True
        else:
            return False
    else:
        return False

def main(xls_path, folder_path, period, captcha_val, window):
    global root
    root = window
    global dwnld_path
    dwnld_path = folder_path
    global path
    path = xls_path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    total_rows = sheet.max_row
    # Initialize the WebDriver with the options
    browser = webdriver.Chrome(options=options)
    browser.maximize_window()
    wait = WebDriverWait(browser, 7)
    action = ActionChains(browser)
    ML.initializeModel()

    row = 2
    execute_finally = True  # Flag to execute the finally block
    try:
        while True:
            try:
            # Iterate through each row in the sheet
                while row <= total_rows:
                    name = sheet.cell(row=row, column=1).value
                    username = sheet.cell(row=row, column=2).value
                    password = sheet.cell(row=row, column=3).value

                    # Set up the download path for each iteration
                    if period == "Quarterly":
                        download_path = f"{folder_path}/{name}-{Quarter}"
                    elif period == "Monthly":
                        download_path = f"{folder_path}/{name}-{Month}"
                    elif period == "All":
                        download_path = f"{folder_path}/{name}-All-({Year})"
                    else:
                        messagebox.showinfo("Error", "Select a valid period to download.")
                        return
                    
                    # Create the download directory if it doesn't exist
                    if not os.path.exists(download_path):
                        os.makedirs(download_path)

                    # Update the download path using CDP command
                    browser.execute_cdp_cmd('Page.setDownloadBehavior', {
                        'behavior': 'allow',
                        'downloadPath': download_path
                    })

                    # Navigate to the GST login page for each iteration
                    browser.get('https://services.gst.gov.in/services/login')
                    sleep(0.5)

                    # Login
                    wait.until(EC.presence_of_element_located((By.NAME, "user_name"))).send_keys(username + Keys.TAB + password + Keys.TAB)
                    # root = tk.Tk()
                    # root.withdraw()
                    # root.after(0, root.lift)
                    # root.focus_force()
                    if captcha_val == "manual":
                        captcha = simpledialog.askstring("Captcha", "Enter the captcha")
                        if captcha is None:
                            browser.quit()
                            return
                    elif captcha_val == "auto":
                        if os.path.exists(f"{dwnld_path}/captcha.png"):
                            os.remove(f"{dwnld_path}/captcha.png")
                        sleep(0.5)
                        element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/form/div[5]/div/div/div/table/tbody/tr[1]/th[1]/img')))
                        element.screenshot(f"{dwnld_path}/captcha.png")
                        captcha = captch_solver(dwnld_path)
                    elif captcha_val == "ML":
                        if os.path.exists(f"{dwnld_path}/captcha.png"):
                            os.remove(f"{dwnld_path}/captcha.png")
                        sleep(0.8)
                        element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div/div/div/div/div/form/div[5]/div/div/div/table/tbody/tr[1]/th[1]/img')))
                        element.screenshot(f"{dwnld_path}/captcha.png")
                        mlpath = f"{dwnld_path}/captcha.png"
                        captcha = ML.inference(mlpath)
                    wait.until(EC.presence_of_element_located((By.NAME, "captcha"))).send_keys(captcha + Keys.ENTER)
                    if os.path.exists(f"{dwnld_path}/captcha.png"):
                        os.remove(f"{dwnld_path}/captcha.png")
                    root.after(1500)

                    # Call captchaerr function to handle incorrect captcha entries
                    if captcha_err(browser, wait, password, captcha_val):
                        return

                    # Call wrong_pass function to handle incorrect password entries
                    if wrong_pass(browser, workbook, row, sheet, download_path):
                        row +=1
                        continue

                    # Call change_pass function to handle change password prompts
                    if change_pass(browser, wait, row, sheet, password, name, workbook):
                        break

                    # Call kyc_auth function to handle KYC prompts
                    kyc_auth(browser, wait)

                    # Continue with the rest of the navigation and downloading process
                    sleep(1.5)
                    services_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/ng-include[2]/nav/div/div/ul/li[2]/a"))).click()
                    returns_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/ng-include[2]/nav/div/div/ul/li[2]/ul/li[4]/a")))
                    action.move_to_element(returns_btn).perform()
                    return_dash_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/ng-include[2]/nav/div/div/ul/li[2]/ul/li[4]/div/ul/li[1]/a"))).click()
                    sleep(1.5)
                    if period == "Quarterly":
                        for i in range(0,3):                            
                            drpdwn_year = wait.until(EC.presence_of_element_located((By.NAME, "fin")))
                            select = Select(drpdwn_year)
                            select.select_by_visible_text(Year)
                            sleep(0.5)
                            drpdwn_quarter = wait.until(EC.presence_of_element_located((By.NAME, "quarter")))
                            select = Select(drpdwn_quarter)
                            select.select_by_visible_text(Quarter)
                            sleep(0.5)
                            drpdwn_month = wait.until(EC.presence_of_element_located((By.NAME, "mon")))
                            select = Select(drpdwn_month)
                            select.select_by_visible_text(Months[i])
                            sleep(0.5)
                            search_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[4]/button"))).click()
                            sleep(0.8)
                            scroll_amount = page_height / 4
                            browser.execute_script(f"window.scrollBy(0, {scroll_amount})")    
                            dwnld_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[3]/div/div/div/div[2]/button | /html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[2]/div/div/div/div[2]/button"))).click()
                            sleep(0.8)
                            dwnld_btn_json = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[5]/button"))).click()
                            
                            sleep(0.5)
                            page_height = browser.execute_script("return document.body.scrollHeight")
                            scroll_amount = page_height / 4
                            browser.execute_script(f"window.scrollBy(0, {scroll_amount})")
                            dwnld_btn_excel = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[6]/button"))).click()
                            browser.execute_script("window.scrollBy(0, document.body.scrollHeight)")
                            sleep(0.5)
                            if i != 2:
                                browser.back()
                                sleep(1.2)
                            if delete_folder_if_empty(download_path):
                                sheet.cell(row=row, column=4).value = "❌"
                                sheet.cell(row=row, column=5).value = "not generated"
                                workbook.save(path)
                            else:
                                sheet.cell(row=row, column=4).value = "✅"
                                workbook.save(path)
                        row +=1
                    elif period == "Monthly":
                        drpdwn_year = wait.until(EC.presence_of_element_located((By.NAME, "fin")))
                        select = Select(drpdwn_year)
                        select.select_by_visible_text(Year)
                        sleep(0.5)
                        drpdwn_quarter = wait.until(EC.presence_of_element_located((By.NAME, "quarter")))
                        select = Select(drpdwn_quarter)
                        select.select_by_visible_text(monthly_quarter)
                        sleep(0.5)
                        drpdwn_month = wait.until(EC.presence_of_element_located((By.NAME, "mon")))
                        select = Select(drpdwn_month)
                        select.select_by_visible_text(Month)
                        sleep(0.5)
                        search_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[4]/button"))).click()
                        browser.execute_script("window.scrollBy(0, document.body.scrollHeight)")
                        dwnld_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[3]/div/div/div/div[2]/button | /html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[2]/div/div/div/div[2]/button"))).click()
                        sleep(0.8)
                        dwnld_btn_json = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[5]/button"))).click()
                        sleep(0.5)
                        page_height = browser.execute_script("return document.body.scrollHeight")
                        scroll_amount = page_height / 4
                        browser.execute_script(f"window.scrollBy(0, {scroll_amount})")
                        dwnld_btn_excel = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[6]/button"))).click()
                        browser.execute_script("window.scrollBy(0, document.body.scrollHeight)")
                        sleep(0.8)
                        if delete_folder_if_empty(download_path):
                            sheet.cell(row=row, column=4).value = "❌"
                            sheet.cell(row=row, column=5).value = "not generated"
                            workbook.save(path)
                        else:
                            sheet.cell(row=row, column=4).value = "✅"
                            workbook.save(path)
                        row +=1
                    elif period == "All":
                        for i in range(0,12):
                            drpdwn_year = wait.until(EC.presence_of_element_located((By.NAME, "fin")))
                            select = Select(drpdwn_year)
                            select.select_by_visible_text(Year)
                            sleep(0.5)
                            drpdwn_quarter = wait.until(EC.presence_of_element_located((By.NAME, "quarter")))
                            select = Select(drpdwn_quarter)
                            quarter_decide(i)
                            select.select_by_visible_text(all_quarter)
                            sleep(0.5)
                            drpdwn_month = wait.until(EC.presence_of_element_located((By.NAME, "mon")))
                            select = Select(drpdwn_month)
                            select.select_by_visible_text(all_months[i])
                            sleep(0.5)
                            search_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[4]/button"))).click()
                            sleep(0.8)
                            browser.execute_script("window.scrollBy(0, document.body.scrollHeight)")
                            dwnld_btn = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[3]/div/div/div/div[2]/button | /html/body/div[2]/div[2]/div/div[2]/div[4]/div[3]/div[1]/div[2]/div/div/div/div[2]/button"))).click()
                            sleep(0.8)
                            dwnld_btn_json = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[5]/button"))).click()
                            sleep(0.5)
                            page_height = browser.execute_script("return document.body.scrollHeight")
                            scroll_amount = page_height / 4
                            browser.execute_script(f"window.scrollBy(0, {scroll_amount})")
                            dwnld_btn_excel = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/div[2]/app-gstr2bdwld/block-ui/div/div[2]/div/div[6]/button"))).click()
                            browser.execute_script("window.scrollBy(0, document.body.scrollHeight)")
                            sleep(0.5)
                            browser.back()
                            sleep(0.8)
                            if delete_folder_if_empty(download_path):
                                sheet.cell(row=row, column=4).value = "❌"
                                sheet.cell(row=row, column=5).value = "not generated"
                                workbook.save(path)
                            else:
                                sheet.cell(row=row, column=4).value = "✅"
                                workbook.save(path)
                        row +=1
                    else: 
                        messagebox.showinfo("Error", "Select a valid period to download.")
            except TimeoutException:
                # Refresh the page if a timeout exception occurs
                browser.refresh()
                continue
            except NoSuchElementException as e:
                # Use a regular expression to extract the message content
                match = re.search(r'Message: (.+?);', str(e))
                if match:
                    message = match.group(1)
                else:
                    message = str(e)  # fallback if pattern not found
                messagebox.showinfo("Exception", f"An exception occurred: {message}")
                execute_finally = False 
                break
            except Exception as e:
                match = re.search(r'Message: (.+?);', str(e))
                if match:
                    message = match.group(1)
                else:
                    message = str(e)  # fallback if pattern not found
                # Handle any other exceptions
                messagebox.showinfo("Exception", f"An exception occurred: {message}")
                if os.path.exists(f"{dwnld_path}/captcha.png"):
                    os.remove(f"{dwnld_path}/captcha.png")
                workbook.save(path)
                try:
                    workbook.close()
                except:
                    pass
                try:
                    browser.quit()
                except:
                    pass
                execute_finally = False # Set the flag to False to prevent the finally block from executing
                break
    finally:
        if execute_finally:        
            # Close the workbook and quit the browser
            messagebox.showinfo("Info", "All files have been downloaded." )
            if os.path.exists(f"{dwnld_path}/captcha.png"):
                os.remove(f"{dwnld_path}/captcha.png")
            try:
                workbook.close()
            except:
                pass
            try:
                browser.quit()
            except:
                pass